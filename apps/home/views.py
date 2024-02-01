from django.http import JsonResponse
from django.http import HttpResponse
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.timezone import localtime
from .models import *
from .forms import *
import openpyxl


def test(request):
    context = {'segment':'dashboard'}
    return render(request, 'home/samples/icon-feather.html', context)#Carga de plantilla de ejemplos

@login_required(login_url="/login/")
def dashboard(request):
    productos_list = Producto.objects.all().order_by('-ventas')[:5]#Listar los 5 productos mas vendidos
    #Listar a todos los ordenes
    ordenes_list = Orden.objects.all().order_by('-fecha_orden')[:5]

    context = {'productos':productos_list,'ordenes':ordenes_list, 'segment':'dashboard'}
    return render(request, 'home/dashboard.html', context)#Cargar la platilla del dashboard

#Clientes----------------------------------------------------------------------------------------->
@login_required(login_url="/login/")
@permission_required('home.view_cliente', raise_exception=True)#Validar permiso
def clientes(request):
    clientes_list = Cliente.objects.all().order_by('id')#Listar a todos los clientes

    texto_busqueda = request.GET.get('texto_busqueda')#Tomar texto del buscador
    if texto_busqueda:#Si exste, filtrar
        clientes_list = clientes_list.filter(Q(nombre__icontains=texto_busqueda)|Q(apellido__icontains=texto_busqueda)|Q(telefono__icontains=texto_busqueda)|Q(estatus__icontains=texto_busqueda))#Filtros
    
    paginator = Paginator(clientes_list, per_page=13, orphans=2)#Iniciar el paginador
    page_number = request.GET.get('page')#Obtener el numero de página
    clientes = paginator.get_page(page_number)#Devolver la lista de objetos paginada

    table = 'home/table-segments/clientes.html'
    
    context = {'items':clientes, 'segment':'clientes', 'title':'Clientes', 'table': table}#Diccionario de objetos que se pasarán a la platilla

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':#Evaluar si es una petición AJAX
        table_html = render_to_string(table, context, request)#Rendereziar los datos en una plantilla de tabla reducida
        paginator_html = render_to_string('includes/paginator.html', context, request)#Rendereziar el paginador actualizado
        return JsonResponse({'table_html': table_html, 'paginator_html': paginator_html})#JsonResponse para manejar con JavaScript y recargar un segmento de la página

    return render(request, 'home/tables.html', context)#Renderizar la plantilla normalmente con los datos del diccionario

@login_required(login_url="/login/")
@permission_required('home.add_cliente', raise_exception=True)#Validar permiso
def clienteCrear(request):
    
    form = ClienteForm(request.POST or None)#Llamar al formulario y cargarle datos si existen (si hay un error en el formulario y la pàgina recarga)

    if form.is_valid():#Validar formulario
        form.save()#Guardar
        return redirect('clientes')#Redireccionar

    context = {'form':form, 'segment': 'clientes', 'titulo':'Agregar Cliente'}#Diccionario de objetos que se pasarán a la platilla
    return render(request, 'home/form/cliente.html', context)#Renderizar la plantilla normalmente con los datos del diccionario
    
@login_required(login_url="/login/")
@permission_required('home.change_cliente', raise_exception=True)#Validar permiso
def clienteEditar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)#Obtener el cliente a editar
    form = ClienteForm(request.POST or None, instance=cliente)#Llamar al formulario y cargarle datos del cliente

    if form.is_valid():#Validar formulario
        form.save()#Guardar
        return redirect('clientes')#Redireccionar

    context = {'form':form, 'segment': 'clientes', 'titulo':'Editar Cliente'}#Diccionario de objetos que se pasarán a la platilla
    return render(request, 'home/form/cliente.html', context)#Renderizar la plantilla normalmente con los datos del diccionario

@login_required(login_url="/login/")
@permission_required('home.delete_cliente', raise_exception=True)#Validar permiso
def clienteEliminar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)#Obtener el cliente a eliminar
    cliente.delete()#Eliminar

    if 'next' in request.GET:
        return redirect(request.GET.get('next'))#Evaluar si existe una página a la que redireccionar y redireccionar
    return redirect('clientes')#Redireccionar normalmente
#Fin Clientes------------------------------------------------------------------------------------->

#Prdouctos---------------------------------------------------------------------------------------->
@login_required(login_url="/login/")
@permission_required('home.view_producto', raise_exception=True)#Validar permiso
def productos(request):
    productos_list = Producto.objects.all().order_by('id')#Listar a todos los productos

    texto_busqueda = request.GET.get('texto_busqueda')#Tomar texto del buscador
    if texto_busqueda:#Si exste, filtrar
        productos_list = productos_list.filter(Q(nombre__icontains=texto_busqueda)|Q(tipo__icontains=texto_busqueda)|Q(codigo__icontains=texto_busqueda)|Q(stock__icontains=texto_busqueda))#Filtros
    
    paginator = Paginator(productos_list, per_page=13, orphans=2)#Iniciar el paginador
    page_number = request.GET.get('page')#Obtener el numero de página
    productos = paginator.get_page(page_number)#Devolver la lista de objetos paginada

    table ='home/table-segments/productos.html'
    
    context = {'items':productos, 'segment':'productos', 'title':'Productos', 'table':table }#Diccionario de objetos que se pasarán a la platilla

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':#Evaluar si es una petición AJAX
        table_html = render_to_string(table, context, request)#Rendereziar los datos en una plantilla de tabla reducida
        paginator_html = render_to_string('includes/paginator.html', context, request)#Rendereziar el paginador actualizado
        return JsonResponse({'table_html': table_html, 'paginator_html': paginator_html})#JsonResponse para manejar con JavaScript y recargar un segmento de la página

    return render(request, 'home/tables.html', context)#Renderizar la plantilla normalmente con los datos del diccionario

@login_required(login_url="/login/")
@permission_required('home.add_producto', raise_exception=True)#Validar permiso
def productoCrear(request):
    
    form = ProductoForm(request.POST or None)#Llamar al formulario y cargarle datos si existen (si hay un error en el formulario y la pàgina recarga)

    if form.is_valid():#Validar formulario
        form.save()#Guardar
        return redirect('productos')#Redireccionar

    context = {'form':form, 'segment': 'productos', 'titulo':'Agregar Producto'}#Diccionario de objetos que se pasarán a la platilla
    return render(request, 'home/form/producto.html', context)#Renderizar la plantilla normalmente con los datos del diccionario
    
@login_required(login_url="/login/")
@permission_required('home.change_producto', raise_exception=True)#Validar permiso
def productoEditar(request, pk):
    producto = get_object_or_404(Producto, pk=pk)#Obtener el producto a editar
    form = ProductoForm(request.POST or None, instance=producto)#Llamar al formulario y cargarle datos del producto

    if form.is_valid():#Validar formulario
        form.save()#Guardar
        return redirect('productos')#Redireccionar

    context = {'form':form, 'segment': 'productos', 'titulo':'Editar Producto'}#Diccionario de objetos que se pasarán a la platilla
    return render(request, 'home/form/producto.html', context)#Renderizar la plantilla normalmente con los datos del diccionario

@login_required(login_url="/login/")
@permission_required('home.delete_producto', raise_exception=True)#Validar permiso
def productoEliminar(request, pk):
    producto = get_object_or_404(Producto, pk=pk)#Obtener el producto a eliminar
    producto.delete()#Eliminar

    if 'next' in request.GET:
        return redirect(request.GET.get('next'))#Evaluar si existe una página a la que redireccionar y redireccionar
    return redirect('productos')#Redireccionar normalmente
#Fin Productos------------------------------------------------------------------------------------>

#Órdenes------------------------------------------------------------------------------------------>
@login_required(login_url="/login/")
@permission_required('home.view_orden', raise_exception=True)#Validar permiso
def ordenes(request):
    ordenes_list = Orden.objects.all().order_by('id')#Listar a todos los ordenes

    texto_busqueda = request.GET.get('texto_busqueda')#Tomar texto del buscador
    if texto_busqueda:#Si exste, filtrar
        ordenes_list = ordenes_list.filter(Q(cliente__nombre__icontains=texto_busqueda)|Q(cliente__apellido__icontains=texto_busqueda)|Q(fecha_orden__icontains=texto_busqueda)|Q(cliente__ci__icontains=texto_busqueda)|Q(completada__icontains=texto_busqueda))#Filtros
    
    paginator = Paginator(ordenes_list, per_page=13, orphans=2)#Iniciar el paginador
    page_number = request.GET.get('page')#Obtener el numero de página
    ordenes = paginator.get_page(page_number)#Devolver la lista de objetos paginada

    table = 'home/table-segments/ordenes.html'

    context = {'items':ordenes, 'segment':'ordenes', 'title':'Órdenes de compra', 'table': table}#Diccionario de objetos que se pasarán a la platilla

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':#Evaluar si es una petición AJAX
        table_html = render_to_string(table, context, request)#Rendereziar los datos en una plantilla de tabla reducida
        paginator_html = render_to_string('includes/paginator.html', context, request)#Rendereziar el paginador actualizado
        return JsonResponse({'table_html': table_html, 'paginator_html': paginator_html})#JsonResponse para manejar con JavaScript y recargar un segmento de la página

    return render(request, 'home/tables.html', context)#Renderizar la plantilla normalmente con los datos del diccionario

@login_required(login_url="/login/")
@permission_required('home.add_orden', raise_exception=True)#Validar permiso
def ordenCrear(request):
    
    form = OrdenForm(request.POST or None)#Llamar al formulario y cargarle datos si existen (si hay un error en el formulario y la pàgina recarga)
    formset = OrdenItemFormset(request.POST or None, prefix='orden_item')#Llamar al formulario hijo y cargarle datos si existen (si hay un error en el formulario y la pàgina recarga)

    if form.is_valid() and formset.is_valid():#Validar formulario
        orden = form.save(commit=False)#Instanciar objeto en la base de datos y asignar a una variable para manipularlo antes de guardar
        #Espacio par aplicar validaciones de backend
        
        orden.save()

        items = formset.save(commit=False)#Instanciar objeto en la base de datos y asignar a una variable para manipularlo antes de guardar
        for orden_item in items:#Recorrer la lista de objetos
            orden_item.orden = orden#asignar elemento hijo al padre
            #Espacio par aplicar validaciones de backend
            orden_item.save()#guardar elemento hijo
            
        return redirect('ordenes')#Redireccionar

    context = {'form':form, 'formset':formset, 'segment': 'ordenes', 'titulo':'Crear Orden de venta'}#Diccionario de objetos que se pasarán a la platilla
    return render(request, 'home/form/orden.html', context)#Renderizar la plantilla normalmente con los datos del diccionario
    
@login_required(login_url="/login/")
@permission_required('home.change_orden', raise_exception=True)#Validar permiso
def ordenEditar(request, pk):
    orden = get_object_or_404(Orden, pk=pk)#Obtener la orden a editar

    form = OrdenForm(request.POST or None, instance=orden)#Llamar al formulario y cargarle datos de la orden
    formset = OrdenItemFormset(request.POST or None, instance=orden, prefix='orden_item')#Llamar al formulario hijo y cargarle datos si existen (si hay un error en el formulario y la pàgina recarga)

    if form.is_valid():#Validar formulario
        form.save()#Guardar
        return redirect('ordenes')#Redireccionar

    context = {'form':form, 'formset':formset, 'segment': 'ordenes', 'titulo':'Editar Orden de venta'}#Diccionario de objetos que se pasarán a la platilla
    return render(request, 'home/form/orden.html', context)#Renderizar la plantilla normalmente con los datos del diccionario

@login_required(login_url="/login/")
@permission_required('home.delete_orden', raise_exception=True)#Validar permiso
def ordenEliminar(request, pk):
    orden = get_object_or_404(Orden, pk=pk)#Obtener la orden a eliminar
    orden.delete()#Eliminar

    if 'next' in request.GET:
        return redirect(request.GET.get('next'))#Evaluar si existe una página a la que redireccionar y redireccionar
    return redirect('ordenes')#Redireccionar normalmente
#Fin Órdenes-------------------------------------------------------------------------------------->

#reportes------------------------------------------------------------------------------------------>
@login_required(login_url="/login/")
@permission_required('home.view_orden', raise_exception=True)
def reporteOrdenes(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Añade los títulos de los campos en la primera fila
    titles = ['ID', 'Cliente', 'Fecha de Orden', 'Monto a Pagar', 'Monto Cancelado', 'Estatus']
    ws.append(titles)

    # Filtra las órdenes por el valor de 'estatus'
    estatus = request.GET.get('estatus', '')
    if estatus:
        records = Orden.objects.filter(completada=estatus)
    else:
        records = Orden.objects.all()

    for record in records:
        # Asegúrate de que el objeto de fecha/hora no tenga información de zona horaria
        fecha_orden = localtime(record.fecha_orden).replace(tzinfo=None)

        # Construye la fila con los datos
        row = [
            record.id,
            record.cliente.nombre + ' ' + record.cliente.apellido if record.cliente else '', # Accede al nombre y apellido del cliente
            fecha_orden.strftime('%Y-%m-%d %H:%M:%S'), # Formatea la fecha
            record.monto_pagar,
            record.monto_cacelado,
            record.completada
        ]

        ws.append(row)

    wb.save(response)
    return response


#Reporte de salida de productos
@login_required(login_url="/login/")
@permission_required('home.view_orden', raise_exception=True)#Validar permiso
def productoSalida(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Encabezados para el reporte
    titles = ['Orden de Compra (ID)', 'Fecha de salida', 'Código', 'Nombre', 'Cantidad']
    ws.append(titles)

    # Obtener todas las OrdenItem y sus respectivos Productos
    orden_items = OrdenItem.objects.select_related('producto').all()

    for item in orden_items:
        # Asegúrate de que el objeto de fecha/hora no tenga información de zona horaria
        fecha_salida = localtime(item.fecha_salida).replace(tzinfo=None) if item.fecha_salida else None
        fecha_salida_str = fecha_salida.strftime('%Y-%m-%d %H:%M:%S') if fecha_salida else ''

        # Construye la fila con los datos
        row = [
            item.orden.id if item.orden else '',
            fecha_salida_str,
            item.producto.codigo if item.producto else '',
            item.producto.nombre if item.producto else '',
            item.cantidad
        ]

        ws.append(row)

    wb.save(response)
    return response

#reporte de inventario
@login_required(login_url="/login/")
@permission_required('home.view_producto', raise_exception=True)
def reporteInventario(request):
    # Crear una respuesta HTTP con el contenido adecuado para un archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Inventario.xlsx"'

    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Productos"

    # Agregar los títulos de las columnas
    column_titles = ['Código de Producto', 'Tipo de Producto', 'Nombre', 'Stock', 'Precio', 'Descuento']
    ws.append(column_titles)

    # Consultar todos los productos en la base de datos
    productos = Producto.objects.all()

    # Iterar sobre cada producto y escribir la información en el archivo Excel
    for producto in productos:
        row = [
            producto.codigo,
            producto.tipo,
            producto.nombre,
            producto.stock,
            producto.precio,
            producto.descuento
        ]
        ws.append(row)

    # Guardar el libro de Excel en la respuesta HTTP
    wb.save(response)
    return response

#reporte de clientes
@login_required(login_url="/login/")
@permission_required('home.view_cliente', raise_exception=True)
def reporteClientes(request):
    # Crear una respuesta HTTP con el contenido adecuado para un archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Clientes.xlsx"'

    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Agregar los títulos de las columnas
    column_titles = [ 'Nacionalidad','CI', 'Nombre', 'Apellido', 'Teléfono', 'Dirección', 'Email', 'Estatus']
    ws.append(column_titles)

    # Filtra los clientes por el valor de 'estatus'
    estatus = request.GET.get('estatus', '')
    if estatus:
        clientes = Cliente.objects.filter(estatus=estatus)
    else:
        clientes = Cliente.objects.all()

    # Iterar sobre cada cliente y escribir la información en el archivo Excel
    for cliente in clientes:
        row = [ 
            cliente.get_ci_tipo_display(), # Obtiene la representación de la opción seleccionada
            cliente.ci,
            cliente.nombre,
            cliente.apellido,
            cliente.telefono,
            cliente.direccion,
            cliente.email if cliente.email else '', # Comprueba si hay un email para evitar incluir None
            cliente.get_estatus_display() # Obtiene la representación de la opción seleccionada
        ]
        ws.append(row)

    # Guardar el libro de Excel en la respuesta HTTP
    wb.save(response)
    return response